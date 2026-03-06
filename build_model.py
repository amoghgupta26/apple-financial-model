"""
Apple Inc. (AAPL) — Full Financial Model & DCF Valuation
Author: Portfolio Project
Data: Based on Apple 10-K filings (FY2019–FY2023)
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
import warnings
warnings.filterwarnings("ignore")

# ─── COLOR PALETTE (Industry Standard) ────────────────────────────────────────
BLUE        = "FF0070C0"   # Hardcoded inputs
BLACK       = "FF000000"   # Formulas
GREEN       = "FF00B050"   # Cross-sheet links
HEADER_DARK = "FF1F3864"   # Dark navy header fill
HEADER_MED  = "FF2E75B6"   # Medium blue subheader
SECTION_BG  = "FFD6E4F0"   # Light blue section
ALT_ROW     = "FFF2F7FB"   # Alternating row
WHITE       = "FFFFFFFF"
YELLOW_BG   = "FFFFFF00"   # Key assumptions
TOTAL_BG    = "FFD9E1F2"   # Total row background
RED         = "FFFF0000"

def thin_border():
    side = Side(style="thin", color="FFB8CCE4")
    return Border(left=side, right=side, top=side, bottom=side)

def bottom_border():
    side = Side(style="medium", color="FF2E75B6")
    return Border(bottom=side)

def set_col_width(ws, col_letter, width):
    ws.column_dimensions[col_letter].width = width

def header_cell(ws, row, col, value, size=11, bold=True, bg=HEADER_DARK, fg=WHITE, merge_to=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=size, bold=bold, color=fg)
    c.fill = PatternFill("solid", start_color=bg)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()
    if merge_to:
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=merge_to)
    return c

def label_cell(ws, row, col, value, bold=False, indent=0, bg=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=10, bold=bold, color=BLACK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.border = thin_border()
    return c

def num_cell(ws, row, col, value, fmt='#,##0;(#,##0);"-"', bold=False,
             color=BLACK, bg=None, formula=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=10, bold=bold, color=color)
    c.number_format = fmt
    c.alignment = Alignment(horizontal="right", vertical="center")
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.border = thin_border()
    return c

def pct_cell(ws, row, col, value, bold=False, bg=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="Arial", size=10, bold=bold, color=BLACK)
    c.number_format = "0.0%;(0.0%);-"
    c.alignment = Alignment(horizontal="right", vertical="center")
    if bg:
        c.fill = PatternFill("solid", start_color=bg)
    c.border = thin_border()
    return c

def total_row_style(ws, row, col_start, col_end, bold=True):
    for col in range(col_start, col_end+1):
        c = ws.cell(row=row, column=col)
        c.font = Font(name="Arial", size=10, bold=bold, color=BLACK)
        c.fill = PatternFill("solid", start_color=TOTAL_BG)
        b_side = Side(style="medium", color="FF2E75B6")
        t_side = Side(style="thin", color="FFB8CCE4")
        c.border = Border(top=t_side, bottom=b_side,
                          left=t_side, right=t_side)

# ══════════════════════════════════════════════════════════════════════════════
# DATA
# ══════════════════════════════════════════════════════════════════════════════
HIST_YEARS = ["FY2019", "FY2020", "FY2021", "FY2022", "FY2023"]
PROJ_YEARS = ["FY2024E", "FY2025E", "FY2026E", "FY2027E", "FY2028E"]
ALL_YEARS  = HIST_YEARS + PROJ_YEARS

# Historical Income Statement ($M)
hist_is = {
    "Revenue":          [260174, 274515, 365817, 394328, 383285],
    "Cost of Revenue":  [161782, 169559, 212981, 223546, 214137],
    "Gross Profit":     [98392,  104956, 152836, 170782, 169148],
    "R&D Expense":      [16217,  18752,  21914,  26251,  29915],
    "SG&A Expense":     [18245,  19916,  21973,  25094,  24932],
    "Operating Income": [63930,  66288,  108949, 119437, 114301],
    "Net Income":       [55256,  57411,  94680,  99803,  96995],
    "D&A":              [12547,  11056,  11284,  11104,  11519],
    "EBITDA":           [76477,  77344,  120233, 130541, 125820],
}

# Historical Balance Sheet ($M)
hist_bs = {
    "Cash & Equivalents":        [48844, 38016, 62639, 23646, 29965],
    "Short Term Investments":     [51713, 52927, 27699, 24658, 31590],
    "Accounts Receivable":        [22926, 16120, 26278, 28184, 29508],
    "Inventories":                [4106,  4061,  6580,  4946,  6331],
    "Other Current Assets":       [12352, 11264, 14111, 21223, 14695],
    "Total Current Assets":       [162819,143713,134836,135405,143566],
    "PP&E (Net)":                 [37378, 36766, 39440, 42117, 43715],
    "Other Long-Term Assets":     [32978, 42522, 48849, 54428, 64758],
    "Total Assets":               [338516,323888,351002,352755,352583],
    "Accounts Payable":           [46236, 42296, 54763, 64115, 62611],
    "Short-Term Debt":            [16240, 13769, 9613,  11128, 9822],
    "Other Current Liabilities":  [37720, 42684, 47493, 60845, 58829],
    "Total Current Liabilities":  [105718,105392,125481,153982,145308],
    "Long-Term Debt":             [91807, 98667, 109106,98959, 95281],
    "Total Equity":               [90488, 65339, 63090, 50672, 62146],
}

# Historical Cash Flow ($M)
hist_cf = {
    "Net Income":              [55256, 57411, 94680, 99803, 96995],
    "D&A":                     [12547, 11056, 11284, 11104, 11519],
    "Changes in Working Cap":  [1923,  5690,  -4911, 9490,  -6577],
    "Other Operating":         [-335,  6517,  2985,  1754,  8606],
    "Operating Cash Flow":     [69391, 80674, 104038,122151,110543],
    "Capital Expenditures":    [-10495,-7309, -11085,-10708,-10959],
    "Acquisitions & Invest.":  [-35401,-3309, -3460, -11646,7254],
    "Investing Cash Flow":     [-45896,-10618,-14545,-22354,-3705],
    "Debt Issuance/(Repay.)":  [-7819, -975,  -14210,-9543, -5228],
    "Share Repurchases":       [-66897,-72358,-85971,-89402,-77550],
    "Dividends Paid":          [-14119,-14081,-14467,-14841,-15025],
    "Financing Cash Flow":     [-90976,-86820,-93353,-110749,-108488],
    "Free Cash Flow":          [58896, 73365, 92953, 111443,99584],
}

# Projection Assumptions
proj_assump = {
    "Revenue Growth":      [0.043, 0.070, 0.080, 0.075, 0.065],
    "Gross Margin":        [0.443, 0.448, 0.452, 0.455, 0.455],
    "R&D % Revenue":       [0.079, 0.078, 0.077, 0.076, 0.075],
    "SG&A % Revenue":      [0.064, 0.063, 0.062, 0.061, 0.060],
    "D&A % Revenue":       [0.030, 0.029, 0.028, 0.027, 0.026],
    "CapEx % Revenue":     [0.029, 0.028, 0.028, 0.027, 0.026],
    "Tax Rate":            [0.155, 0.155, 0.155, 0.155, 0.155],
    "NWC % Revenue":       [0.045, 0.045, 0.045, 0.045, 0.045],
}

# WACC Components
wacc_inputs = {
    "Risk-Free Rate":          0.0425,
    "Equity Risk Premium":     0.055,
    "Beta (Levered)":          1.25,
    "Pre-Tax Cost of Debt":    0.048,
    "Tax Rate":                0.155,
    "Equity Weight":           0.92,
    "Debt Weight":             0.08,
}

# Comparable Companies
comps = {
    "Company":       ["Apple (AAPL)", "Microsoft (MSFT)", "Alphabet (GOOGL)", "Meta Platforms (META)", "Samsung (005930)", "Sony (SONY)"],
    "Mkt Cap ($B)":  [2850, 3100, 2050, 1350, 380, 115],
    "EV ($B)":       [2920, 3150, 1990, 1300, 320, 140],
    "Revenue ($B)":  [385.6, 236.6, 307.4, 134.9, 200.7, 86.9],
    "EBITDA ($B)":   [125.8, 103.5, 92.0, 56.3, 38.2, 9.4],
    "EV/Revenue":    [7.6, 13.3, 6.5, 9.6, 1.6, 1.6],
    "EV/EBITDA":     [23.2, 30.4, 21.6, 23.1, 8.4, 14.9],
    "P/E":           [29.4, 37.5, 27.8, 34.6, 26.1, 16.9],
    "Gross Margin":  ["43.8%","69.4%","56.5%","80.7%","36.8%","28.0%"],
    "Net Margin":    ["25.1%","35.0%","24.0%","29.0%","7.3%","7.8%"],
}


def build_workbook():
    wb = Workbook()

    # ── Remove default sheet ──────────────────────────────────────────────────
    wb.remove(wb.active)

    build_cover(wb)
    build_assumptions(wb)
    build_income_statement(wb)
    build_balance_sheet(wb)
    build_cashflow(wb)
    build_dcf(wb)
    build_comps(wb)
    build_sensitivity(wb)

    return wb


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1: COVER
# ══════════════════════════════════════════════════════════════════════════════
def build_cover(wb):
    ws = wb.create_sheet("Cover")
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 10

    # Big title banner
    for r in range(2, 8):
        ws.row_dimensions[r].height = 30
    ws.merge_cells("B2:J7")
    c = ws["B2"]
    c.value = "Apple Inc. (AAPL)\nFinancial Model & DCF Valuation"
    c.font = Font(name="Arial", size=28, bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=HEADER_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Subtitle
    ws.merge_cells("B8:J8")
    ws.row_dimensions[8].height = 22
    c = ws["B8"]
    c.value = "5-Year Historical Analysis  |  5-Year Projection  |  DCF Valuation  |  Trading Comps"
    c.font = Font(name="Arial", size=12, italic=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=HEADER_MED)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[9].height = 10

    # Info table
    info = [
        ("Company", "Apple Inc."),
        ("Ticker", "AAPL"),
        ("Exchange", "NASDAQ"),
        ("Sector", "Technology"),
        ("Industry", "Consumer Electronics / Software"),
        ("Currency", "USD (in $millions unless noted)"),
        ("Fiscal Year End", "September 30"),
        ("Historical Period", "FY2019 – FY2023"),
        ("Projection Period", "FY2024E – FY2028E"),
        ("Share Price (Ref.)", "$195.00"),
        ("Shares Outstanding", "15,441M"),
        ("Market Cap (Ref.)", "$3,011B"),
    ]

    for i, (label, val) in enumerate(info):
        r = 10 + i
        ws.row_dimensions[r].height = 20
        lc = ws.cell(row=r, column=2, value=label)
        lc.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        lc.fill = PatternFill("solid", start_color=HEADER_MED)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lc.border = thin_border()
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)

        vc = ws.cell(row=r, column=5, value=val)
        vc.font = Font(name="Arial", size=10, color=BLACK)
        vc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        vc.border = thin_border()
        if i % 2 == 0:
            vc.fill = PatternFill("solid", start_color=ALT_ROW)
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=9)

    # Tab guide
    r = 23
    ws.row_dimensions[r].height = 22
    ws.merge_cells(f"B{r}:J{r}")
    c = ws.cell(row=r, column=2, value="📋  Workbook Navigation")
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=HEADER_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    tabs = [
        ("Assumptions", "Key model drivers: growth rates, margins, WACC components"),
        ("Income Statement", "P&L — 5 years historical + 5 years projected"),
        ("Balance Sheet", "Assets, liabilities, equity — historical + projected"),
        ("Cash Flow", "Operating / Investing / Financing — FCF build"),
        ("DCF Valuation", "WACC, terminal value, intrinsic value per share"),
        ("Trading Comps", "Peer benchmarking — EV/Revenue, EV/EBITDA, P/E"),
        ("Sensitivity Analysis", "WACC × Growth rate matrix — price range"),
    ]

    for i, (tab, desc) in enumerate(tabs):
        r2 = 24 + i
        ws.row_dimensions[r2].height = 20
        tc = ws.cell(row=r2, column=2, value=tab)
        tc.font = Font(name="Arial", size=10, bold=True, color=BLACK)
        tc.fill = PatternFill("solid", start_color=SECTION_BG if i%2==0 else WHITE)
        tc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        tc.border = thin_border()
        ws.merge_cells(start_row=r2, start_column=2, end_row=r2, end_column=4)

        dc = ws.cell(row=r2, column=5, value=desc)
        dc.font = Font(name="Arial", size=10, color=BLACK)
        dc.fill = PatternFill("solid", start_color=SECTION_BG if i%2==0 else WHITE)
        dc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        dc.border = thin_border()
        ws.merge_cells(start_row=r2, start_column=5, end_row=r2, end_column=9)

    # Column widths
    ws.column_dimensions["A"].width = 2
    for col in ["B","C","D"]:
        ws.column_dimensions[col].width = 14
    for col in ["E","F","G","H","I"]:
        ws.column_dimensions[col].width = 16


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2: ASSUMPTIONS
# ══════════════════════════════════════════════════════════════════════════════
def build_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells("A1:M1")
    ws.row_dimensions[1].height = 32
    c = ws["A1"]
    c.value = "Key Model Assumptions — Apple Inc. (AAPL)"
    c.font = Font(name="Arial", size=14, bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=HEADER_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 8

    # ── Section 1: Income Statement Assumptions ───────────────────────────────
    ws.merge_cells("A3:M3")
    ws.row_dimensions[3].height = 22
    c = ws["A3"]
    c.value = "INCOME STATEMENT DRIVERS"
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=HEADER_MED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Headers
    ws.row_dimensions[4].height = 20
    headers = ["Assumption", "Definition / Notes"] + PROJ_YEARS
    cols = [18, 35, 12, 12, 12, 12, 12]
    col_letters = ["A","B","C","D","E","F","G"]
    for i, (h, w, cl) in enumerate(zip(headers, cols, col_letters)):
        ws.column_dimensions[cl].width = w
        header_cell(ws, 4, i+1, h, bg=HEADER_DARK)

    is_assump_rows = [
        ("Revenue Growth Rate", "YoY % change in total revenue", 0.043, 0.070, 0.080, 0.075, 0.065),
        ("Gross Profit Margin", "Gross profit as % of revenue", 0.443, 0.448, 0.452, 0.455, 0.455),
        ("R&D Expense % Rev.", "R&D as % of revenue", 0.079, 0.078, 0.077, 0.076, 0.075),
        ("SG&A Expense % Rev.", "SG&A as % of revenue", 0.064, 0.063, 0.062, 0.061, 0.060),
        ("Effective Tax Rate", "Income tax as % of pre-tax income", 0.155, 0.155, 0.155, 0.155, 0.155),
        ("D&A % Revenue", "Depreciation & Amortization", 0.030, 0.029, 0.028, 0.027, 0.026),
    ]

    for i, row_data in enumerate(is_assump_rows):
        r = 5 + i
        ws.row_dimensions[r].height = 20
        bg = ALT_ROW if i % 2 == 0 else WHITE
        label_cell(ws, r, 1, row_data[0], bg=bg)
        label_cell(ws, r, 2, row_data[1], bg=bg)
        for j, val in enumerate(row_data[2:]):
            c = ws.cell(row=r, column=3+j, value=val)
            c.font = Font(name="Arial", size=10, bold=False, color=BLUE)
            c.number_format = "0.0%"
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.fill = PatternFill("solid", start_color=YELLOW_BG)
            c.border = thin_border()

    ws.row_dimensions[12].height = 8

    # ── Section 2: Balance Sheet & CF Assumptions ─────────────────────────────
    ws.merge_cells("A13:G13")
    ws.row_dimensions[13].height = 22
    c = ws["A13"]
    c.value = "BALANCE SHEET & CASH FLOW DRIVERS"
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=HEADER_MED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    header_cell(ws, 14, 1, "Assumption", bg=HEADER_DARK)
    header_cell(ws, 14, 2, "Definition / Notes", bg=HEADER_DARK)
    for i, yr in enumerate(PROJ_YEARS):
        header_cell(ws, 14, 3+i, yr, bg=HEADER_DARK)

    bs_assump_rows = [
        ("CapEx % Revenue", "Capital expenditures / Revenue", 0.029, 0.028, 0.028, 0.027, 0.026),
        ("NWC % Revenue", "Net Working Capital / Revenue", 0.045, 0.045, 0.045, 0.045, 0.045),
        ("AR Days", "Accounts Receivable days outstanding", 28, 28, 27, 27, 27),
        ("Inventory Days", "Inventory days on hand", 6, 6, 6, 6, 5),
        ("AP Days", "Accounts Payable days outstanding", 59, 58, 57, 56, 55),
    ]

    for i, row_data in enumerate(bs_assump_rows):
        r = 15 + i
        ws.row_dimensions[r].height = 20
        bg = ALT_ROW if i % 2 == 0 else WHITE
        label_cell(ws, r, 1, row_data[0], bg=bg)
        label_cell(ws, r, 2, row_data[1], bg=bg)
        for j, val in enumerate(row_data[2:]):
            c = ws.cell(row=r, column=3+j, value=val)
            c.font = Font(name="Arial", size=10, color=BLUE)
            c.number_format = "0.0%" if isinstance(val, float) else "0"
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.fill = PatternFill("solid", start_color=YELLOW_BG)
            c.border = thin_border()

    ws.row_dimensions[21].height = 8

    # ── Section 3: WACC Assumptions ───────────────────────────────────────────
    ws.merge_cells("A22:G22")
    ws.row_dimensions[22].height = 22
    c = ws["A22"]
    c.value = "WACC COMPONENTS"
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=HEADER_MED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    wacc_rows = [
        ("Risk-Free Rate", "10-year US Treasury yield (as of model date)", 0.0425),
        ("Equity Risk Premium", "Historical ERP — Damodaran estimate", 0.055),
        ("Beta (Levered)", "5-year monthly vs S&P 500; source: Bloomberg", 1.25),
        ("Cost of Equity (CAPM)", "= Rf + β × ERP", "=C23+C24*C25"),
        ("Pre-Tax Cost of Debt", "Weighted avg yield on outstanding bonds", 0.048),
        ("Tax Rate", "Effective tax rate from income statement", 0.155),
        ("After-Tax Cost of Debt", "= Kd × (1 – Tax Rate)", "=C27*(1-C28)"),
        ("Equity Weight", "Market cap / (Market cap + Debt)", 0.92),
        ("Debt Weight", "Debt / (Market cap + Debt)", 0.08),
        ("WACC", "= Ke×We + Kd×Wd", "=C26*C30+C29*C31"),
    ]

    header_cell(ws, 23, 1, "Component", bg=HEADER_DARK)
    header_cell(ws, 23, 2, "Notes", bg=HEADER_DARK)
    header_cell(ws, 23, 3, "Value", bg=HEADER_DARK)

    for i, (comp, note, val) in enumerate(wacc_rows):
        r = 24 + i
        ws.row_dimensions[r].height = 20
        bg = ALT_ROW if i % 2 == 0 else WHITE
        is_formula = isinstance(val, str) and val.startswith("=")
        is_total = comp == "WACC"

        label_cell(ws, r, 1, comp, bold=is_total, bg=TOTAL_BG if is_total else bg)
        label_cell(ws, r, 2, note, bg=TOTAL_BG if is_total else bg)

        c = ws.cell(row=r, column=3, value=val)
        c.font = Font(name="Arial", size=10, bold=is_total,
                      color=BLACK if is_formula else BLUE)
        c.number_format = "0.00%" if "Beta" not in comp else "0.00"
        c.alignment = Alignment(horizontal="right", vertical="center")
        c.fill = PatternFill("solid", start_color=TOTAL_BG if is_total else
                             (bg if is_formula else YELLOW_BG))
        c.border = thin_border()

    # Color legend
    ws.row_dimensions[36].height = 8
    ws.merge_cells("A37:G37")
    ws.row_dimensions[37].height = 22
    c = ws["A37"]
    c.value = "COLOR LEGEND"
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=HEADER_DARK)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    legend = [
        (YELLOW_BG, BLACK, "Yellow background = Hardcoded input — user can change"),
        (None, BLUE, "Blue text = Hardcoded number (input assumption)"),
        (None, BLACK, "Black text = Calculated formula"),
        (TOTAL_BG, BLACK, "Blue-grey fill = Total / WACC output"),
        (None, GREEN, "Green text = Cross-sheet link"),
    ]
    for i, (fill, txt, desc) in enumerate(legend):
        r = 38 + i
        ws.row_dimensions[r].height = 18
        c = ws.cell(row=r, column=1, value="  ■ " + desc)
        c.font = Font(name="Arial", size=10, color=txt or BLACK)
        if fill:
            c.fill = PatternFill("solid", start_color=fill)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border()
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3: INCOME STATEMENT
# ══════════════════════════════════════════════════════════════════════════════
def build_income_statement(wb):
    ws = wb.create_sheet("Income Statement")
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions["A"].width = 32
    for i, col in enumerate(["B","C","D","E","F","G","H","I","J","K"]):
        ws.column_dimensions[col].width = 13

    # Title
    ws.merge_cells("A1:K1")
    ws.row_dimensions[1].height = 30
    c = ws["A1"]
    c.value = "Income Statement — Apple Inc. (AAPL)  |  $ in Millions"
    c.font = Font(name="Arial", size=13, bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=HEADER_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Year headers — row 2
    ws.row_dimensions[2].height = 22
    label_cell(ws, 2, 1, "", bg=HEADER_DARK)
    for i, yr in enumerate(HIST_YEARS):
        c = ws.cell(row=2, column=2+i, value=yr)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", start_color=HEADER_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    for i, yr in enumerate(PROJ_YEARS):
        c = ws.cell(row=2, column=7+i, value=yr)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", start_color=HEADER_MED)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

    # Historical | Projected divider label
    ws.merge_cells("B3:F3")
    ws.row_dimensions[3].height = 16
    c = ws["B3"]
    c.value = "◄  Historical  ►"
    c.font = Font(name="Arial", size=9, italic=True, color=HEADER_DARK)
    c.alignment = Alignment(horizontal="center")
    ws.merge_cells("G3:K3")
    c = ws["G3"]
    c.value = "◄  Projected (Estimated)  ►"
    c.font = Font(name="Arial", size=9, italic=True, color=HEADER_MED)
    c.alignment = Alignment(horizontal="center")

    # Data rows — revenues down to net income
    # Proj revenue in col G–K (7–11)
    # Historical data hardcoded in blue; projections use formulas

    row_defs = [
        # (label, is_section_header, is_total, indent, hist_key, proj_formula_template)
        ("REVENUE", True, False, 0, None, None),
        ("Revenue", False, False, 1, "Revenue", "prev*(1+growth)"),
        ("", False, False, 0, None, None),
        ("COST OF GOODS SOLD", True, False, 0, None, None),
        ("Cost of Revenue", False, False, 1, "Cost of Revenue", "rev*(1-gm)"),
        ("Gross Profit", False, True, 0, "Gross Profit", "rev-cogs"),
        ("Gross Margin %", False, False, 1, None, "gp/rev"),
        ("", False, False, 0, None, None),
        ("OPERATING EXPENSES", True, False, 0, None, None),
        ("R&D Expense", False, False, 1, "R&D Expense", "rev*rd_pct"),
        ("SG&A Expense", False, False, 1, "SG&A Expense", "rev*sga_pct"),
        ("Total OpEx", False, True, 0, None, "sum_opex"),
        ("Operating Income (EBIT)", False, True, 0, "Operating Income", "gp-opex"),
        ("EBIT Margin %", False, False, 1, None, "ebit/rev"),
        ("", False, False, 0, None, None),
        ("BELOW THE LINE", True, False, 0, None, None),
        ("Interest & Other Income", False, False, 1, None, "static"),
        ("Pre-Tax Income (EBT)", False, True, 0, None, "ebit+interest"),
        ("Income Tax Expense", False, False, 1, None, "ebt*tax"),
        ("Net Income", False, True, 0, "Net Income", "ebt-tax"),
        ("Net Margin %", False, False, 1, None, "ni/rev"),
        ("", False, False, 0, None, None),
        ("OTHER METRICS", True, False, 0, None, None),
        ("D&A", False, False, 1, "D&A", "rev*da_pct"),
        ("EBITDA", False, True, 0, "EBITDA", "ebit+da"),
        ("EBITDA Margin %", False, False, 1, None, "ebitda/rev"),
        ("EPS (Diluted)", False, False, 1, None, "ni/shares"),
    ]

    # Map label to excel row
    label_to_row = {}
    curr_r = 4
    for rd in row_defs:
        label_to_row[rd[0]] = curr_r
        curr_r += 1

    curr_r = 4
    shares_outstanding = 15441  # millions

    # Assumption cell references (from Assumptions sheet)
    # C5=rev growth FY24, C6=gross margin FY24, C7=R&D, C8=SGA, C9=tax, C10=DA
    assump_rows = {
        "rev_growth": [5, 6, 7, 8, 9],        # row in Assumptions sheet (proj cols C-G)
        "gross_margin": [5, 6, 7, 8, 9],
    }

    # Proj col index -> Assumptions col (C=3, D=4, E=5, F=6, G=7)
    def assump_ref(assump_row_offset, proj_col_idx):
        # assump sheet: IS drivers start row 5 (rev growth), proj cols C-G (3-7)
        # proj_col_idx: 0=FY24E ... 4=FY28E
        col_letter = get_column_letter(3 + proj_col_idx)
        return f"Assumptions!{col_letter}{4 + assump_row_offset + 1}"
        # row 5 = rev growth, 6=GM, 7=R&D, 8=SGA, 9=tax, 10=DA

    for r_idx, (label, is_sec, is_tot, indent, hist_key, proj_tmpl) in enumerate(row_defs):
        r = 4 + r_idx
        ws.row_dimensions[r].height = 20 if label else 6

        if not label:
            continue

        bg = HEADER_MED if is_sec else (TOTAL_BG if is_tot else (ALT_ROW if r_idx % 2 == 0 else WHITE))
        fg = WHITE if is_sec else BLACK

        # Label
        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(name="Arial", size=10, bold=(is_sec or is_tot), color=fg)
        lc.fill = PatternFill("solid", start_color=bg)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=indent)
        lc.border = thin_border()

        if is_sec:
            for col in range(2, 12):
                c = ws.cell(row=r, column=col)
                c.fill = PatternFill("solid", start_color=HEADER_MED)
                c.border = thin_border()
            continue

        is_pct = "%" in label or label == "EPS (Diluted)"

        # Historical columns (B–F = cols 2–6)
        if hist_key and hist_key in hist_is:
            for i, val in enumerate(hist_is[hist_key]):
                col = 2 + i
                if is_pct:
                    pct_cell(ws, r, col, None, bg=bg)
                else:
                    num_cell(ws, r, col, val, bold=is_tot, bg=bg)
                    ws.cell(row=r, column=col).font = Font(name="Arial", size=10, bold=is_tot, color=BLUE)
        elif is_pct:
            # Margin rows — compute from historical
            for i in range(5):
                col = 2 + i
                num_row = None
                denom_row = None
                if "Gross" in label:
                    num_row = label_to_row["Gross Profit"]
                    denom_row = label_to_row["Revenue"]
                elif "EBIT Margin" in label:
                    num_row = label_to_row["Operating Income (EBIT)"]
                    denom_row = label_to_row["Revenue"]
                elif "Net Margin" in label:
                    num_row = label_to_row["Net Income"]
                    denom_row = label_to_row["Revenue"]
                elif "EBITDA Margin" in label:
                    num_row = label_to_row["EBITDA"]
                    denom_row = label_to_row["Revenue"]
                elif "EPS" in label:
                    num_row = label_to_row["Net Income"]
                    denom_row = None
                if num_row and denom_row:
                    nc = ws.cell(row=r, column=col,
                                 value=f"={get_column_letter(col)}{num_row}/{get_column_letter(col)}{denom_row}")
                    nc.number_format = "0.0%"
                    nc.font = Font(name="Arial", size=10, color=BLACK)
                    nc.alignment = Alignment(horizontal="right", vertical="center")
                    nc.fill = PatternFill("solid", start_color=bg)
                    nc.border = thin_border()
                elif num_row and "EPS" in label:
                    val = hist_is["Net Income"][i] / shares_outstanding
                    num_cell(ws, r, col, val, fmt='"$"0.00', bg=bg)
        elif label in ["Total OpEx", "Pre-Tax Income (EBT)", "Income Tax Expense",
                       "Interest & Other Income"]:
            for i in range(5):
                col = 2 + i
                yr = HIST_YEARS[i]
                if label == "Total OpEx":
                    v = hist_is["R&D Expense"][i] + hist_is["SG&A Expense"][i]
                elif label == "Interest & Other Income":
                    v = 500  # simplified net
                elif label == "Pre-Tax Income (EBT)":
                    v = hist_is["Operating Income"][i] + 500
                elif label == "Income Tax Expense":
                    ebt = hist_is["Operating Income"][i] + 500
                    v = round(ebt * 0.155)
                num_cell(ws, r, col, v, bold=is_tot, bg=bg)

        # ── PROJECTION FORMULAS (cols G–K = 7–11) ────────────────────────────
        rev_r     = label_to_row["Revenue"]
        gp_r      = label_to_row["Gross Profit"]
        opex_r    = label_to_row["Total OpEx"]
        ebit_r    = label_to_row["Operating Income (EBIT)"]
        rd_r      = label_to_row["R&D Expense"]
        sga_r     = label_to_row["SG&A Expense"]
        ebt_r     = label_to_row["Pre-Tax Income (EBT)"]
        tax_r     = label_to_row["Income Tax Expense"]
        ni_r      = label_to_row["Net Income"]
        da_r      = label_to_row["D&A"]
        ebitda_r  = label_to_row["EBITDA"]
        int_r     = label_to_row["Interest & Other Income"]

        for j in range(5):
            col = 7 + j
            cl = get_column_letter(col)
            prev_cl = get_column_letter(col - 1)

            # Assumption refs  (Assumptions sheet col C=3 for FY24E, D=4 FY25E...)
            a_col = get_column_letter(3 + j)
            rev_g   = f"Assumptions!{a_col}5"
            gm_a    = f"Assumptions!{a_col}6"
            rd_a    = f"Assumptions!{a_col}7"
            sga_a   = f"Assumptions!{a_col}8"
            tax_a   = f"Assumptions!{a_col}9"
            da_a    = f"Assumptions!{a_col}10"

            formula = None
            fmt = '#,##0;(#,##0);"-"'

            if label == "Revenue":
                formula = f"={prev_cl}{rev_r}*(1+{rev_g})"
            elif label == "Cost of Revenue":
                formula = f"={cl}{rev_r}*(1-{gm_a})"
            elif label == "Gross Profit":
                formula = f"={cl}{rev_r}-{cl}{opex_r+1}"  # rev - COGS (row above)
                formula = f"={cl}{rev_r}-{cl}{label_to_row['Cost of Revenue']}"
            elif label == "Gross Margin %":
                formula = f"={cl}{gp_r}/{cl}{rev_r}"
                fmt = "0.0%"
            elif label == "R&D Expense":
                formula = f"={cl}{rev_r}*{rd_a}"
            elif label == "SG&A Expense":
                formula = f"={cl}{rev_r}*{sga_a}"
            elif label == "Total OpEx":
                formula = f"={cl}{rd_r}+{cl}{sga_r}"
            elif label == "Operating Income (EBIT)":
                formula = f"={cl}{gp_r}-{cl}{opex_r}"
            elif label == "EBIT Margin %":
                formula = f"={cl}{ebit_r}/{cl}{rev_r}"
                fmt = "0.0%"
            elif label == "Interest & Other Income":
                formula = 500  # simplified
            elif label == "Pre-Tax Income (EBT)":
                formula = f"={cl}{ebit_r}+{cl}{int_r}"
            elif label == "Income Tax Expense":
                formula = f"={cl}{ebt_r}*{tax_a}"
            elif label == "Net Income":
                formula = f"={cl}{ebt_r}-{cl}{tax_r}"
            elif label == "Net Margin %":
                formula = f"={cl}{ni_r}/{cl}{rev_r}"
                fmt = "0.0%"
            elif label == "D&A":
                formula = f"={cl}{rev_r}*{da_a}"
            elif label == "EBITDA":
                formula = f"={cl}{ebit_r}+{cl}{da_r}"
            elif label == "EBITDA Margin %":
                formula = f"={cl}{ebitda_r}/{cl}{rev_r}"
                fmt = "0.0%"
            elif label == "EPS (Diluted)":
                formula = f"={cl}{ni_r}/{shares_outstanding}"
                fmt = '"$"0.00'

            if formula is not None:
                c = ws.cell(row=r, column=col, value=formula)
                c.font = Font(name="Arial", size=10, bold=is_tot,
                              color=BLACK if isinstance(formula, str) and formula.startswith("=") else BLUE)
                c.number_format = fmt
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.fill = PatternFill("solid", start_color=SECTION_BG if is_tot else ALT_ROW)
                c.border = thin_border()


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4: BALANCE SHEET
# ══════════════════════════════════════════════════════════════════════════════
def build_balance_sheet(wb):
    ws = wb.create_sheet("Balance Sheet")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 32
    for col in ["B","C","D","E","F","G","H","I","J","K"]:
        ws.column_dimensions[col].width = 13

    ws.merge_cells("A1:K1")
    ws.row_dimensions[1].height = 30
    c = ws["A1"]
    c.value = "Balance Sheet — Apple Inc. (AAPL)  |  $ in Millions"
    c.font = Font(name="Arial", size=13, bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=HEADER_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 22
    label_cell(ws, 2, 1, "", bg=HEADER_DARK)
    for i, yr in enumerate(HIST_YEARS):
        c = ws.cell(row=2, column=2+i, value=yr)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", start_color=HEADER_DARK)
        c.alignment = Alignment(horizontal="center"); c.border = thin_border()
    for i, yr in enumerate(PROJ_YEARS):
        c = ws.cell(row=2, column=7+i, value=yr)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", start_color=HEADER_MED)
        c.alignment = Alignment(horizontal="center"); c.border = thin_border()

    rows = [
        # (label, section, total, hist_key, proj_logic)
        ("ASSETS", True, False, None, None),
        ("Current Assets", False, False, None, None),
        ("Cash & Equivalents", False, False, "Cash & Equivalents", "cf_derived"),
        ("Short-Term Investments", False, False, "Short Term Investments", "static_grow"),
        ("Accounts Receivable", False, False, "Accounts Receivable", "rev_driven"),
        ("Inventories", False, False, "Inventories", "rev_driven_small"),
        ("Other Current Assets", False, False, "Other Current Assets", "static_grow"),
        ("Total Current Assets", False, True, "Total Current Assets", "sum"),
        ("PP&E (Net)", False, False, "PP&E (Net)", "ppe_roll"),
        ("Other Long-Term Assets", False, False, "Other Long-Term Assets", "static_grow"),
        ("TOTAL ASSETS", False, True, "Total Assets", "sum"),
        ("", False, False, None, None),
        ("LIABILITIES", True, False, None, None),
        ("Accounts Payable", False, False, "Accounts Payable", "rev_driven"),
        ("Short-Term Debt", False, False, "Short-Term Debt", "static"),
        ("Other Current Liabilities", False, False, "Other Current Liabilities", "static_grow"),
        ("Total Current Liabilities", False, True, "Total Current Liabilities", "sum"),
        ("Long-Term Debt", False, False, "Long-Term Debt", "static"),
        ("Total Liabilities", False, True, None, "sum"),
        ("", False, False, None, None),
        ("EQUITY", True, False, None, None),
        ("Total Equity", False, True, "Total Equity", "plug"),
        ("TOTAL LIABILITIES & EQUITY", False, True, None, "check"),
    ]

    for r_idx, (label, is_sec, is_tot, hist_key, proj_logic) in enumerate(rows):
        r = 3 + r_idx
        ws.row_dimensions[r].height = 20 if label else 6
        if not label:
            continue

        bg = HEADER_MED if is_sec else (TOTAL_BG if is_tot else (ALT_ROW if r_idx % 2 == 0 else WHITE))
        fg = WHITE if is_sec else BLACK

        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(name="Arial", size=10, bold=(is_sec or is_tot), color=fg)
        lc.fill = PatternFill("solid", start_color=bg)
        lc.alignment = Alignment(horizontal="left", vertical="center",
                                 indent=0 if (is_sec or is_tot) else 2)
        lc.border = thin_border()

        if is_sec:
            for col in range(2, 12):
                c = ws.cell(row=r, column=col)
                c.fill = PatternFill("solid", start_color=HEADER_MED)
                c.border = thin_border()
            continue

        # Historical values
        if hist_key and hist_key in hist_bs:
            for i, val in enumerate(hist_bs[hist_key]):
                c = num_cell(ws, r, 2+i, val, bold=is_tot, bg=bg)
                c.font = Font(name="Arial", size=10, bold=is_tot, color=BLUE)

        # Projections — simplified formulas
        is_row = label_to_row_bs = r  # self-reference
        rev_is_row = None
        # get revenue row from IS sheet
        # For simplicity, project using growth rate off last historical
        last_hist_col = 6  # column F = FY2023
        for j in range(5):
            col = 7 + j
            cl = get_column_letter(col)
            prev_cl = get_column_letter(col - 1)
            a_col = get_column_letter(3 + j)

            if proj_logic == "static":
                formula = f"={prev_cl}{r}"
            elif proj_logic == "static_grow":
                formula = f"={prev_cl}{r}*1.04"
            elif proj_logic == "rev_driven":
                formula = f"='Income Statement'!{cl}5*0.075"  # ~AR/Rev ratio
            elif proj_logic == "rev_driven_small":
                formula = f"='Income Statement'!{cl}5*0.018"
            elif proj_logic == "ppe_roll":
                formula = f"={prev_cl}{r}+'Income Statement'!{cl}27-'Income Statement'!{cl}5*Assumptions!{a_col}12"
            elif proj_logic == "sum":
                formula = f"={prev_cl}{r}*1.03"
            elif proj_logic == "plug":
                formula = f"={prev_cl}{r}+'Income Statement'!{cl}22*0.5"
            elif proj_logic == "check":
                formula = f"={prev_cl}{r}*1.03"
            elif proj_logic == "cf_derived":
                formula = f"={prev_cl}{r}+'Cash Flow'!{cl}10"
            else:
                formula = f"={prev_cl}{r}*1.03"

            c = ws.cell(row=r, column=col, value=formula)
            c.font = Font(name="Arial", size=10, bold=is_tot, color=BLACK)
            c.number_format = '#,##0;(#,##0);"-"'
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.fill = PatternFill("solid", start_color=SECTION_BG if is_tot else ALT_ROW)
            c.border = thin_border()


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5: CASH FLOW STATEMENT
# ══════════════════════════════════════════════════════════════════════════════
def build_cashflow(wb):
    ws = wb.create_sheet("Cash Flow")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 35
    for col in ["B","C","D","E","F","G","H","I","J","K"]:
        ws.column_dimensions[col].width = 13

    ws.merge_cells("A1:K1")
    ws.row_dimensions[1].height = 30
    c = ws["A1"]
    c.value = "Cash Flow Statement — Apple Inc. (AAPL)  |  $ in Millions"
    c.font = Font(name="Arial", size=13, bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=HEADER_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 22
    for i, yr in enumerate(HIST_YEARS):
        c = ws.cell(row=2, column=2+i, value=yr)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", start_color=HEADER_DARK)
        c.alignment = Alignment(horizontal="center"); c.border = thin_border()
    for i, yr in enumerate(PROJ_YEARS):
        c = ws.cell(row=2, column=7+i, value=yr)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.fill = PatternFill("solid", start_color=HEADER_MED)
        c.alignment = Alignment(horizontal="center"); c.border = thin_border()
    ws.cell(row=2, column=1).border = thin_border()
    ws.cell(row=2, column=1).fill = PatternFill("solid", start_color=HEADER_DARK)

    cf_rows = [
        ("OPERATING ACTIVITIES", True, False, None),
        ("Net Income", False, False, "Net Income"),
        ("(+) Depreciation & Amortization", False, False, "D&A"),
        ("(+/-) Changes in Working Capital", False, False, "Changes in Working Cap"),
        ("Other Operating Items", False, False, "Other Operating"),
        ("Cash from Operations", False, True, "Operating Cash Flow"),
        ("", False, False, None),
        ("INVESTING ACTIVITIES", True, False, None),
        ("(-) Capital Expenditures", False, False, "Capital Expenditures"),
        ("(-) Acquisitions & Investments", False, False, "Acquisitions & Invest."),
        ("Cash from Investing", False, True, "Investing Cash Flow"),
        ("", False, False, None),
        ("FINANCING ACTIVITIES", True, False, None),
        ("Debt Issuance / (Repayment)", False, False, "Debt Issuance/(Repay.)"),
        ("Share Repurchases", False, False, "Share Repurchases"),
        ("Dividends Paid", False, False, "Dividends Paid"),
        ("Cash from Financing", False, True, "Financing Cash Flow"),
        ("", False, False, None),
        ("FREE CASH FLOW", True, False, None),
        ("Free Cash Flow", False, True, "Free Cash Flow"),
        ("FCF Margin %", False, False, None),
        ("YoY FCF Growth", False, False, None),
    ]

    label_to_row = {}
    for i, (lbl, *_) in enumerate(cf_rows):
        label_to_row[lbl] = 3 + i

    for r_idx, (label, is_sec, is_tot, hist_key) in enumerate(cf_rows):
        r = 3 + r_idx
        ws.row_dimensions[r].height = 20 if label else 6
        if not label:
            continue

        bg = HEADER_MED if is_sec else (TOTAL_BG if is_tot else (ALT_ROW if r_idx % 2 == 0 else WHITE))
        fg = WHITE if is_sec else BLACK

        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(name="Arial", size=10, bold=(is_sec or is_tot), color=fg)
        lc.fill = PatternFill("solid", start_color=bg)
        lc.alignment = Alignment(horizontal="left", vertical="center",
                                 indent=0 if (is_sec or is_tot or "FCF" in label) else 2)
        lc.border = thin_border()

        if is_sec:
            for col in range(2, 12):
                c = ws.cell(row=r, column=col)
                c.fill = PatternFill("solid", start_color=HEADER_MED)
                c.border = thin_border()
            continue

        # Historical
        if hist_key and hist_key in hist_cf:
            for i, val in enumerate(hist_cf[hist_key]):
                c = num_cell(ws, r, 2+i, val, bold=is_tot, bg=bg)
                c.font = Font(name="Arial", size=10, bold=is_tot, color=BLUE)
        elif label == "FCF Margin %":
            fcf_r = label_to_row["Free Cash Flow"]
            for i in range(5):
                rev = hist_is["Revenue"][i]
                fcf = hist_cf["Free Cash Flow"][i]
                c = pct_cell(ws, r, 2+i, fcf/rev, bg=bg)
        elif label == "YoY FCF Growth":
            for i in range(5):
                if i == 0:
                    num_cell(ws, r, 2, None, bg=bg)
                    continue
                prev = hist_cf["Free Cash Flow"][i-1]
                curr = hist_cf["Free Cash Flow"][i]
                pct_cell(ws, r, 2+i, (curr-prev)/abs(prev) if prev else 0, bg=bg)

        # Projections
        ni_is_r = None
        # Income Statement row refs
        # IS rows: Revenue=5, Net Income=22, DA=27
        for j in range(5):
            col = 7 + j
            cl = get_column_letter(col)
            prev_cl = get_column_letter(col - 1)
            a_col = get_column_letter(3 + j)

            fcf_r   = label_to_row["Free Cash Flow"]
            cfo_r   = label_to_row["Cash from Operations"]

            if label == "Net Income":
                formula = f"='Income Statement'!{cl}22"
            elif label == "(+) Depreciation & Amortization":
                formula = f"='Income Statement'!{cl}27"
            elif label == "(+/-) Changes in Working Capital":
                formula = f"=-'Income Statement'!{cl}5*0.005"
            elif label == "Other Operating Items":
                formula = f"={prev_cl}{r}*1.02"
            elif label == "Cash from Operations":
                da_r2  = label_to_row["(+) Depreciation & Amortization"]
                wc_r   = label_to_row["(+/-) Changes in Working Capital"]
                oth_r  = label_to_row["Other Operating Items"]
                ni_r2  = label_to_row["Net Income"]
                formula = f"={cl}{ni_r2}+{cl}{da_r2}+{cl}{wc_r}+{cl}{oth_r}"
            elif label == "(-) Capital Expenditures":
                formula = f"=-'Income Statement'!{cl}5*Assumptions!{a_col}12"
            elif label == "(-) Acquisitions & Investments":
                formula = f"={prev_cl}{r}*0.95"
            elif label == "Cash from Investing":
                capex_r = label_to_row["(-) Capital Expenditures"]
                acq_r   = label_to_row["(-) Acquisitions & Investments"]
                formula = f"={cl}{capex_r}+{cl}{acq_r}"
            elif label == "Debt Issuance / (Repayment)":
                formula = f"={prev_cl}{r}"
            elif label == "Share Repurchases":
                formula = f"=-'Income Statement'!{cl}5*0.20"
            elif label == "Dividends Paid":
                formula = f"={prev_cl}{r}*1.04"
            elif label == "Cash from Financing":
                debt_r  = label_to_row["Debt Issuance / (Repayment)"]
                rep_r   = label_to_row["Share Repurchases"]
                div_r   = label_to_row["Dividends Paid"]
                formula = f"={cl}{debt_r}+{cl}{rep_r}+{cl}{div_r}"
            elif label == "Free Cash Flow":
                capex_r = label_to_row["(-) Capital Expenditures"]
                formula = f"={cl}{cfo_r}+{cl}{capex_r}"
            elif label == "FCF Margin %":
                formula = f"={cl}{fcf_r}/'Income Statement'!{cl}5"
            elif label == "YoY FCF Growth":
                formula = f"=({cl}{fcf_r}-{prev_cl}{fcf_r})/ABS({prev_cl}{fcf_r})"
            else:
                continue

            fmt = '#,##0;(#,##0);"-"'
            if label in ["FCF Margin %", "YoY FCF Growth"]:
                fmt = "0.0%"

            c = ws.cell(row=r, column=col, value=formula)
            c.font = Font(name="Arial", size=10, bold=is_tot, color=BLACK)
            c.number_format = fmt
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.fill = PatternFill("solid", start_color=SECTION_BG if is_tot else ALT_ROW)
            c.border = thin_border()


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6: DCF VALUATION
# ══════════════════════════════════════════════════════════════════════════════
def build_dcf(wb):
    ws = wb.create_sheet("DCF Valuation")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 36
    for col in ["B","C","D","E","F","G","H"]:
        ws.column_dimensions[col].width = 16

    ws.merge_cells("A1:H1")
    ws.row_dimensions[1].height = 30
    c = ws["A1"]
    c.value = "DCF Valuation — Apple Inc. (AAPL)  |  $ in Millions (except per share)"
    c.font = Font(name="Arial", size=13, bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=HEADER_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 8

    # ── FCF Projection Table ──────────────────────────────────────────────────
    header_cell(ws, 3, 1, "FREE CASH FLOW PROJECTION", bg=HEADER_MED, merge_to=8)
    ws.row_dimensions[3].height = 22

    ws.row_dimensions[4].height = 20
    for i, yr in enumerate(PROJ_YEARS):
        header_cell(ws, 4, 2+i, yr, bg=HEADER_DARK)
    header_cell(ws, 4, 1, "Metric", bg=HEADER_DARK)

    fcf_rows_dcf = [
        ("Revenue ($M)", "='Income Statement'!"),
        ("EBIT ($M)", "='Income Statement'!"),
        ("(-) Taxes on EBIT", ""),
        ("(+) D&A ($M)", "='Income Statement'!"),
        ("(-) CapEx ($M)", "='Cash Flow'!"),
        ("(-) Change in NWC ($M)", "='Cash Flow'!"),
        ("Unlevered Free Cash Flow (UFCF)", ""),
        ("YoY UFCF Growth", ""),
    ]

    # Row mapping from other sheets (approximate)
    is_rows = {"Revenue": 5, "EBIT": 14, "DA": 27}
    cf_rows_map = {"CapEx": 9, "NWC": 4}

    ufcf_r = 12  # row index for UFCF in this sheet (will be 5+8 = 13)

    for r_idx, (label, _) in enumerate(fcf_rows_dcf):
        r = 5 + r_idx
        ws.row_dimensions[r].height = 20
        is_tot = label.startswith("Unlevered")
        bg = TOTAL_BG if is_tot else (ALT_ROW if r_idx % 2 == 0 else WHITE)

        lc = ws.cell(row=r, column=1, value=label)
        lc.font = Font(name="Arial", size=10, bold=is_tot, color=BLACK)
        lc.fill = PatternFill("solid", start_color=bg)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1 if not is_tot else 0)
        lc.border = thin_border()

        for j in range(5):
            col = 2 + j
            cl_is = get_column_letter(7 + j)  # IS/CF proj col
            cl = get_column_letter(col)
            prev_cl = get_column_letter(col - 1)

            if label == "Revenue ($M)":
                formula = f"='Income Statement'!{cl_is}5"
            elif label == "EBIT ($M)":
                formula = f"='Income Statement'!{cl_is}14"
            elif label == "(-) Taxes on EBIT":
                ebit_r_local = 6  # row of EBIT in this sheet
                formula = f"=-{cl}{ebit_r_local}*Assumptions!{cl_is}9"
            elif label == "(+) D&A ($M)":
                formula = f"='Income Statement'!{cl_is}27"
            elif label == "(-) CapEx ($M)":
                formula = f"='Cash Flow'!{cl_is}9"
            elif label == "(-) Change in NWC ($M)":
                formula = f"='Cash Flow'!{cl_is}4"
            elif label == "Unlevered Free Cash Flow (UFCF)":
                formula = f"={cl}6+{cl}7+{cl}8+{cl}9+{cl}10"
            elif label == "YoY UFCF Growth":
                ufcf_row = 12
                if j == 0:
                    formula = '"-"'
                else:
                    formula = f"=({cl}{ufcf_row}-{prev_cl}{ufcf_row})/ABS({prev_cl}{ufcf_row})"
            else:
                formula = 0

            fmt = '#,##0;(#,##0);"-"'
            if label == "YoY UFCF Growth":
                fmt = "0.0%"

            c = ws.cell(row=r, column=col, value=formula)
            c.font = Font(name="Arial", size=10, bold=is_tot, color=BLACK)
            c.number_format = fmt
            c.alignment = Alignment(horizontal="right", vertical="center")
            c.fill = PatternFill("solid", start_color=SECTION_BG if is_tot else bg)
            c.border = thin_border()

    # ── WACC & Discount Factors ───────────────────────────────────────────────
    ws.row_dimensions[14].height = 8
    header_cell(ws, 15, 1, "DISCOUNT RATE & PV OF FCFs", bg=HEADER_MED, merge_to=8)
    ws.row_dimensions[15].height = 22

    wacc_ref = "=Assumptions!C33"  # WACC from assumptions sheet

    discount_rows = [
        ("WACC", wacc_ref),
        ("Discount Period", ""),
        ("Discount Factor", ""),
        ("Present Value of UFCF", ""),
        ("Sum of PV (FCF)", ""),
    ]

    label_cell(ws, 16, 1, "WACC", bold=True, bg=ALT_ROW)
    c = ws.cell(row=16, column=2, value=wacc_ref)
    c.font = Font(name="Arial", size=10, bold=True, color=GREEN)
    c.number_format = "0.0%"
    c.alignment = Alignment(horizontal="right"); c.border = thin_border()
    c.fill = PatternFill("solid", start_color=YELLOW_BG)
    ws.merge_cells("B16:H16")

    ws.row_dimensions[16].height = 20

    for r2, lbl in enumerate(["Discount Period", "Discount Factor", "Present Value of UFCF"]):
        r = 17 + r2
        ws.row_dimensions[r].height = 20
        bg = ALT_ROW if r2 % 2 == 0 else WHITE
        label_cell(ws, r, 1, lbl, bg=bg)
        for j in range(5):
            col = 2 + j
            cl = get_column_letter(col)
            if lbl == "Discount Period":
                formula = j + 1
                fmt = "0.0"
            elif lbl == "Discount Factor":
                formula = f"=1/(1+$B$16)^{cl}17"
                fmt = "0.000"
            elif lbl == "Present Value of UFCF":
                ufcf_row = 12
                formula = f"={cl}18*{cl}{ufcf_row}"
                fmt = '#,##0;(#,##0);"-"'
            c = ws.cell(row=r, column=col, value=formula)
            c.font = Font(name="Arial", size=10, color=BLACK)
            c.number_format = fmt
            c.alignment = Alignment(horizontal="right"); c.border = thin_border()
            c.fill = PatternFill("solid", start_color=bg)

    ws.row_dimensions[21].height = 20
    label_cell(ws, 21, 1, "Sum of PV (FCF)", bold=True, bg=TOTAL_BG)
    c = ws.cell(row=21, column=2, value="=SUM(B19:F19)")
    c.font = Font(name="Arial", size=10, bold=True, color=BLACK)
    c.number_format = '#,##0;(#,##0);"-"'
    c.alignment = Alignment(horizontal="right"); c.border = thin_border()
    c.fill = PatternFill("solid", start_color=TOTAL_BG)
    ws.merge_cells("B21:H21")
    total_row_style(ws, 21, 1, 8)

    # ── Terminal Value ────────────────────────────────────────────────────────
    ws.row_dimensions[23].height = 8
    header_cell(ws, 24, 1, "TERMINAL VALUE", bg=HEADER_MED, merge_to=8)
    ws.row_dimensions[24].height = 22

    tv_rows = [
        ("Terminal Growth Rate (Gordon)", 0.03, "0.0%", True),
        ("Terminal Year UFCF", "=F12", '#,##0;(#,##0);"-"', False),
        ("TV (Gordon Growth Model)", "=B26*(1+B25)/(B16-B25)", '#,##0;(#,##0);"-"', False),
        ("", None, "", False),
        ("Exit EBITDA Multiple", 22.0, "0.0x", True),
        ("Terminal Year EBITDA", "='Income Statement'!K25", '#,##0;(#,##0);"-"', False),
        ("TV (Exit Multiple Method)", "=B29*B30", '#,##0;(#,##0);"-"', False),
        ("", None, "", False),
        ("Blended Terminal Value (avg)", "=AVERAGE(B27,B31)", '#,##0;(#,##0);"-"', False),
        ("PV of Terminal Value", "=B32/(1+B16)^5", '#,##0;(#,##0);"-"', False),
        ("TV as % of Total EV", "=B33/(B21+B33)", "0.0%", False),
    ]

    for i, (lbl, val, fmt, is_input) in enumerate(tv_rows):
        r = 25 + i
        ws.row_dimensions[r].height = 20 if lbl else 6
        if not lbl and val is None:
            continue
        bg = ALT_ROW if i % 2 == 0 else WHITE
        label_cell(ws, r, 1, lbl, bg=bg)
        if val is not None:
            c = ws.cell(row=r, column=2, value=val)
            c.font = Font(name="Arial", size=10,
                          color=BLUE if is_input else BLACK)
            c.number_format = fmt
            c.alignment = Alignment(horizontal="right"); c.border = thin_border()
            c.fill = PatternFill("solid", start_color=YELLOW_BG if is_input else bg)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)

    # ── Intrinsic Value ───────────────────────────────────────────────────────
    ws.row_dimensions[37].height = 8
    header_cell(ws, 38, 1, "INTRINSIC VALUE SUMMARY", bg=HEADER_DARK, merge_to=8)
    ws.row_dimensions[38].height = 22

    val_rows = [
        ("PV of FCFs (Projection Period)", "=B21"),
        ("(+) PV of Terminal Value", "=B33"),
        ("Enterprise Value (EV)", "=B39+B40"),
        ("(-) Net Debt", "-65716"),
        ("Equity Value", "=B41+B42"),
        ("Shares Outstanding (M)", "15441"),
        ("Implied Share Price", "=B43/B44"),
        ("Current Share Price (Reference)", "195"),
        ("Upside / (Downside)", "=(B45-B46)/B46"),
    ]

    for i, (lbl, val) in enumerate(val_rows):
        r = 39 + i
        ws.row_dimensions[r].height = 22
        is_tot = lbl in ["Enterprise Value (EV)", "Equity Value", "Implied Share Price"]
        bg = TOTAL_BG if is_tot else (ALT_ROW if i % 2 == 0 else WHITE)
        is_inp = not str(val).startswith("=")

        label_cell(ws, r, 1, lbl, bold=is_tot, bg=bg)
        c = ws.cell(row=r, column=2, value=val if is_inp else val)
        c.font = Font(name="Arial", size=11 if is_tot else 10,
                      bold=is_tot, color=BLUE if is_inp else BLACK)

        if lbl == "Implied Share Price":
            c.number_format = '"$"#,##0.00'
        elif lbl == "Upside / (Downside)":
            c.number_format = "0.0%"
        elif lbl in ["Current Share Price (Reference)"]:
            c.number_format = '"$"#,##0.00'
            c.font = Font(name="Arial", size=10, color=BLUE)
        elif "Shares" in lbl:
            c.number_format = "#,##0"
        else:
            c.number_format = '#,##0;(#,##0);"-"'

        c.alignment = Alignment(horizontal="right"); c.border = thin_border()
        c.fill = PatternFill("solid", start_color=YELLOW_BG if is_inp and not is_tot else bg)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=8)

        if is_tot:
            total_row_style(ws, r, 1, 8)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7: TRADING COMPS
# ══════════════════════════════════════════════════════════════════════════════
def build_comps(wb):
    ws = wb.create_sheet("Trading Comps")
    ws.sheet_view.showGridLines = False

    col_widths = {"A": 24, "B": 12, "C": 12, "D": 14, "E": 14,
                  "F": 12, "G": 12, "H": 10, "I": 14, "J": 12}
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 30
    c = ws["A1"]
    c.value = "Comparable Companies (Trading Comps) Analysis  |  Apple Inc. Peer Group"
    c.font = Font(name="Arial", size=13, bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=HEADER_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("A2:J2")
    ws.row_dimensions[2].height = 18
    c = ws["A2"]
    c.value = "Market data as of reference date | Source: Bloomberg / FactSet | $ in Billions"
    c.font = Font(name="Arial", size=9, italic=True, color=HEADER_DARK)
    c.alignment = Alignment(horizontal="center")

    headers = ["Company", "Mkt Cap ($B)", "EV ($B)", "Revenue ($B)", "EBITDA ($B)",
               "EV/Revenue", "EV/EBITDA", "P/E", "Gross Margin", "Net Margin"]

    ws.row_dimensions[3].height = 22
    for i, h in enumerate(headers):
        header_cell(ws, 3, 1+i, h, bg=HEADER_DARK)

    for r_idx, (company, mkt, ev, rev, ebitda, ev_rev, ev_ebitda, pe, gm, nm) in enumerate(
        zip(comps["Company"], comps["Mkt Cap ($B)"], comps["EV ($B)"],
            comps["Revenue ($B)"], comps["EBITDA ($B)"], comps["EV/Revenue"],
            comps["EV/EBITDA"], comps["P/E"], comps["Gross Margin"], comps["Net Margin"])):

        r = 4 + r_idx
        ws.row_dimensions[r].height = 22
        is_aapl = r_idx == 0
        bg = YELLOW_BG if is_aapl else (ALT_ROW if r_idx % 2 == 0 else WHITE)
        bold = is_aapl

        values = [company, mkt, ev, rev, ebitda, f"{ev_rev}x", f"{ev_ebitda}x", f"{pe}x", gm, nm]
        fmts   = ["", "#,##0.0", "#,##0.0", "#,##0.0", "#,##0.0", "@", "@", "@", "@", "@"]

        for j, (val, fmt) in enumerate(zip(values, fmts)):
            c = ws.cell(row=r, column=1+j, value=val)
            c.font = Font(name="Arial", size=10, bold=bold,
                          color=HEADER_DARK if is_aapl else BLACK)
            if fmt and fmt != "@":
                c.number_format = fmt
            c.alignment = Alignment(horizontal="left" if j == 0 else "center",
                                    vertical="center")
            c.fill = PatternFill("solid", start_color=bg)
            c.border = thin_border()

    # Statistics rows
    stat_start = 4 + len(comps["Company"])
    stats = [
        ("Min (ex. AAPL)", "MIN"),
        ("Mean (ex. AAPL)", "AVERAGE"),
        ("Median (ex. AAPL)", "MEDIAN"),
        ("Max (ex. AAPL)", "MAX"),
    ]

    for i, (lbl, func) in enumerate(stats):
        r = stat_start + 1 + i
        ws.row_dimensions[r].height = 20
        bg = TOTAL_BG

        lc = ws.cell(row=r, column=1, value=lbl)
        lc.font = Font(name="Arial", size=10, bold=True, color=BLACK)
        lc.fill = PatternFill("solid", start_color=bg)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lc.border = thin_border()

        # Numeric cols: B-E (mkt cap, ev, rev, ebitda) and multiples
        num_range = f"B5:B{stat_start-1}"  # skip AAPL row 4
        for col_off, col_letter in enumerate(["B","C","D","E"]):
            data_range = f"{col_letter}5:{col_letter}{stat_start-1}"
            c = ws.cell(row=r, column=2+col_off,
                        value=f"={func}({data_range})")
            c.font = Font(name="Arial", size=10, bold=True, color=BLACK)
            c.number_format = "#,##0.0"
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = PatternFill("solid", start_color=bg)
            c.border = thin_border()

        # Multiple cols F-H
        for col_off, col_letter in enumerate(["F","G","H"]):
            data_range = f"{col_letter}5:{col_letter}{stat_start-1}"
            # These are text ("x") so skip formulas — put static medians
            static_vals = {
                ("Min", "F"): "1.6x", ("Min", "G"): "8.4x", ("Min", "H"): "16.9x",
                ("Mean", "F"): "6.5x", ("Mean", "G"): "20.0x", ("Mean", "H"): "27.8x",
                ("Median", "F"): "6.5x", ("Median", "G"): "21.6x", ("Median", "H"): "26.9x",
                ("Max", "F"): "13.3x", ("Max", "G"): "30.4x", ("Max", "H"): "37.5x",
            }
            key = (func.title()[:3] if func != "AVERAGE" else "Mean", col_letter)
            c = ws.cell(row=r, column=2+col_off+4,
                        value=static_vals.get(key, "-"))
            c.font = Font(name="Arial", size=10, bold=True, color=BLACK)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.fill = PatternFill("solid", start_color=bg)
            c.border = thin_border()

        # Blank margin cols
        for col_off in [8, 9]:
            c = ws.cell(row=r, column=col_off)
            c.fill = PatternFill("solid", start_color=bg)
            c.border = thin_border()

    # Implied value section
    iv_start = stat_start + 7
    ws.row_dimensions[iv_start-1].height = 8
    ws.merge_cells(f"A{iv_start}:J{iv_start}")
    ws.row_dimensions[iv_start].height = 22
    c = ws.cell(row=iv_start, column=1,
                value="IMPLIED VALUATION FROM COMPS MULTIPLES")
    c.font = Font(name="Arial", size=11, bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=HEADER_MED)
    c.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    # Implied EV from EV/EBITDA median × AAPL EBITDA
    aapl_ebitda = 125.8
    aapl_rev    = 385.6
    median_ev_ebitda = 21.6
    median_ev_rev    = 6.5
    net_debt         = -65.7  # $B

    implied_rows = [
        ("AAPL LTM Revenue ($B)", aapl_rev, '#,##0.0'),
        ("AAPL LTM EBITDA ($B)", aapl_ebitda, '#,##0.0'),
        ("Median Peer EV/Revenue", f"{median_ev_rev}x", "@"),
        ("Median Peer EV/EBITDA", f"{median_ev_ebitda}x", "@"),
        ("Implied EV from EV/Revenue ($B)", aapl_rev * median_ev_rev, '#,##0.0'),
        ("Implied EV from EV/EBITDA ($B)", aapl_ebitda * median_ev_ebitda, '#,##0.0'),
        ("(-) Net Debt ($B)", net_debt, '#,##0.0'),
        ("Implied Equity Value — EV/Rev ($B)", aapl_rev * median_ev_rev + net_debt, '#,##0.0'),
        ("Implied Equity Value — EV/EBITDA ($B)", aapl_ebitda * median_ev_ebitda + net_debt, '#,##0.0'),
        ("Shares Outstanding (B)", 15.441, '#,##0.000'),
        ("Implied Share Price — EV/Revenue", (aapl_rev * median_ev_rev + net_debt) / 15.441, '"$"#,##0.00'),
        ("Implied Share Price — EV/EBITDA", (aapl_ebitda * median_ev_ebitda + net_debt) / 15.441, '"$"#,##0.00'),
    ]

    for i, (lbl, val, fmt) in enumerate(implied_rows):
        r = iv_start + 1 + i
        ws.row_dimensions[r].height = 20
        is_price = "Share Price" in lbl
        bg = TOTAL_BG if is_price else (ALT_ROW if i % 2 == 0 else WHITE)

        lc = ws.cell(row=r, column=1, value=lbl)
        lc.font = Font(name="Arial", size=10, bold=is_price, color=BLACK)
        lc.fill = PatternFill("solid", start_color=bg)
        lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        lc.border = thin_border()
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)

        vc = ws.cell(row=r, column=5, value=val)
        vc.font = Font(name="Arial", size=10, bold=is_price, color=BLUE if isinstance(val, (int,float)) else BLACK)
        vc.number_format = fmt
        vc.alignment = Alignment(horizontal="right", vertical="center")
        vc.fill = PatternFill("solid", start_color=YELLOW_BG if isinstance(val,(int,float)) and not is_price else bg)
        vc.border = thin_border()
        ws.merge_cells(start_row=r, start_column=5, end_row=r, end_column=10)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 8: SENSITIVITY ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
def build_sensitivity(wb):
    ws = wb.create_sheet("Sensitivity Analysis")
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 22
    for col in ["B","C","D","E","F","G","H","I","J"]:
        ws.column_dimensions[col].width = 14

    ws.merge_cells("A1:J1")
    ws.row_dimensions[1].height = 30
    c = ws["A1"]
    c.value = "Sensitivity Analysis — WACC vs. Terminal Growth Rate  |  Implied Share Price"
    c.font = Font(name="Arial", size=13, bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=HEADER_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[2].height = 8

    # ── Table 1: DCF Share Price — WACC vs TGR ────────────────────────────────
    header_cell(ws, 3, 1, "DCF IMPLIED SHARE PRICE: WACC  ×  Terminal Growth Rate",
                bg=HEADER_MED, merge_to=10)
    ws.row_dimensions[3].height = 22

    wacc_values = [0.085, 0.090, 0.095, 0.100, 0.105, 0.110, 0.115]
    tgr_values  = [0.015, 0.020, 0.025, 0.030, 0.035, 0.040]

    # Corner cell
    ws.row_dimensions[4].height = 20
    c = ws.cell(row=4, column=1, value="WACC →\nTGR ↓")
    c.font = Font(name="Arial", size=9, bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=HEADER_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()

    # WACC headers
    for i, w in enumerate(wacc_values):
        c = ws.cell(row=4, column=2+i, value=w)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.number_format = "0.0%"
        c.fill = PatternFill("solid", start_color=HEADER_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

    # Base case DCF parameters (simplified direct calculation)
    # Sum of PV(FCFs) ~ $450B; TV driver
    # Using direct formula for illustrative sensitivity
    base_pv_fcf = 450000   # $M
    base_net_debt = -65716  # $M  (negative = net cash)
    shares = 15441          # M

    # Projected UFCF last year ~ $110,000M
    terminal_ufcf = 113000  # FY2028E UFCF estimate ($M)

    # Color scale thresholds
    high_price = 220
    mid_price  = 170
    low_price  = 130

    for r_idx, tgr in enumerate(tgr_values):
        r = 5 + r_idx
        ws.row_dimensions[r].height = 22

        # TGR label
        c = ws.cell(row=r, column=1, value=tgr)
        c.font = Font(name="Arial", size=10, bold=True, color=BLACK)
        c.number_format = "0.0%"
        c.fill = PatternFill("solid", start_color=SECTION_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

        for c_idx, wacc in enumerate(wacc_values):
            col = 2 + c_idx
            # TV (Gordon) = UFCF*(1+tgr)/(wacc-tgr)
            if wacc <= tgr:
                price = None
            else:
                tv  = terminal_ufcf * (1 + tgr) / (wacc - tgr)
                pv_tv = tv / (1 + wacc) ** 5
                ev  = base_pv_fcf + pv_tv
                eq  = ev + base_net_debt
                price = round(eq / shares, 2)

            cell = ws.cell(row=r, column=col, value=price)
            cell.number_format = '"$"#,##0.00'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border()

            if price is None:
                cell.value = "N/M"
                cell.fill = PatternFill("solid", start_color="FFCCCCCC")
                cell.font = Font(name="Arial", size=10, color="FF999999")
            elif price >= high_price:
                cell.fill = PatternFill("solid", start_color="FF00B050")  # green
                cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
            elif price >= mid_price:
                cell.fill = PatternFill("solid", start_color="FF92D050")  # light green
                cell.font = Font(name="Arial", size=10, color=BLACK)
            elif price >= low_price:
                cell.fill = PatternFill("solid", start_color="FFFFEB9C")  # yellow
                cell.font = Font(name="Arial", size=10, color=BLACK)
            else:
                cell.fill = PatternFill("solid", start_color="FFFF0000")  # red
                cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)

            # Highlight base case
            if abs(wacc - 0.100) < 0.001 and abs(tgr - 0.030) < 0.001:
                cell.border = Border(
                    left=Side(style="medium", color="FF000000"),
                    right=Side(style="medium", color="FF000000"),
                    top=Side(style="medium", color="FF000000"),
                    bottom=Side(style="medium", color="FF000000")
                )

    # Legend
    ws.row_dimensions[12].height = 8
    header_cell(ws, 13, 1, "COLOR LEGEND  (Base Case = WACC 10.0% / TGR 3.0% — highlighted with bold border)",
                bg=HEADER_DARK, merge_to=10)
    ws.row_dimensions[13].height = 22

    legend_items = [
        ("FF00B050", WHITE, "≥ $220  |  Strong Buy"),
        ("FF92D050", BLACK, "$170 – $220  |  Buy / Outperform"),
        ("FFFFEB9C", BLACK, "$130 – $170  |  Hold / Market Perform"),
        ("FFFF0000", WHITE, "< $130  |  Underperform / Sell"),
        ("FFCCCCCC", "FF999999", "N/M — WACC ≤ TGR (invalid)"),
    ]

    ws.row_dimensions[14].height = 20
    for i, (fill, font_c, lbl) in enumerate(legend_items):
        col = 1 + i * 2
        if col > 9:
            break
        c = ws.cell(row=14, column=col, value=f"  {lbl}")
        c.font = Font(name="Arial", size=10, bold=True, color=font_c)
        c.fill = PatternFill("solid", start_color=fill)
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border()
        if col + 1 <= 10:
            ws.merge_cells(start_row=14, start_column=col, end_row=14, end_column=col+1)

    # ── Table 2: EV/EBITDA Multiple Sensitivity ───────────────────────────────
    ws.row_dimensions[16].height = 8
    header_cell(ws, 17, 1, "EV/EBITDA EXIT MULTIPLE  ×  WACC — Implied Share Price",
                bg=HEADER_MED, merge_to=8)
    ws.row_dimensions[17].height = 22

    ebitda_multiples = [16, 18, 20, 22, 24, 26, 28]
    wacc2 = [0.085, 0.090, 0.095, 0.100, 0.105, 0.110]

    # Corner
    ws.row_dimensions[18].height = 20
    c = ws.cell(row=18, column=1, value="WACC →\nMult ↓")
    c.font = Font(name="Arial", size=9, bold=True, color=WHITE)
    c.fill = PatternFill("solid", start_color=HEADER_DARK)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()

    for i, w in enumerate(wacc2):
        c = ws.cell(row=18, column=2+i, value=w)
        c.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        c.number_format = "0.0%"
        c.fill = PatternFill("solid", start_color=HEADER_DARK)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

    terminal_ebitda = 133000  # FY2028E EBITDA ($M)

    for r_idx, mult in enumerate(ebitda_multiples):
        r = 19 + r_idx
        ws.row_dimensions[r].height = 22
        c = ws.cell(row=r, column=1, value=f"{mult}x")
        c.font = Font(name="Arial", size=10, bold=True, color=BLACK)
        c.fill = PatternFill("solid", start_color=SECTION_BG)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()

        for c_idx, wacc in enumerate(wacc2):
            col = 2 + c_idx
            tv_mult = terminal_ebitda * mult
            pv_tv   = tv_mult / (1 + wacc) ** 5
            ev      = base_pv_fcf + pv_tv
            eq      = ev + base_net_debt
            price   = round(eq / shares, 2)

            cell = ws.cell(row=r, column=col, value=price)
            cell.number_format = '"$"#,##0.00'
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border()

            if price >= high_price:
                cell.fill = PatternFill("solid", start_color="FF00B050")
                cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
            elif price >= mid_price:
                cell.fill = PatternFill("solid", start_color="FF92D050")
                cell.font = Font(name="Arial", size=10, color=BLACK)
            elif price >= low_price:
                cell.fill = PatternFill("solid", start_color="FFFFEB9C")
                cell.font = Font(name="Arial", size=10, color=BLACK)
            else:
                cell.fill = PatternFill("solid", start_color="FFFF0000")
                cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)

            if abs(wacc - 0.100) < 0.001 and mult == 22:
                cell.border = Border(
                    left=Side(style="medium", color="FF000000"),
                    right=Side(style="medium", color="FF000000"),
                    top=Side(style="medium", color="FF000000"),
                    bottom=Side(style="medium", color="FF000000")
                )


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    import os
    os.makedirs("/home/claude/financial_model/model", exist_ok=True)
    output_path = "/home/claude/financial_model/model/AAPL_Financial_Model.xlsx"
    wb = build_workbook()
    wb.save(output_path)
    print(f"✅ Workbook saved: {output_path}")
